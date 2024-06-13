from ensurepip import bootstrap
import pandas as pd
import numpy as np
from decouple import config
from dotenv import load_dotenv, dotenv_values
from flask_bootstrap import Bootstrap

import meraki
from meraki import APIError
# The built-in OS module, to read environment variables
import os
# We're also going to import Python's built-in JSON module, but only to make the console output pretty. In production, you wouldn't need any of the printing calls at all, nor this import!
import json

# The openpyxl module, to manipulate Excel files
import openpyxl
# The datetime module, to generate timestamps
import datetime

import threading
import time
import random

from flask import Flask, jsonify, render_template, request, redirect, url_for, session

load_dotenv()

JOB_ORDER_EXCEL_NAME = 'excel.xlsx'
API_KEY = os.getenv('MERAKI_DASHBOARD_API_KEY')
USE_PROXY = os.getenv('USE_PROXY')
PROXY_CERTIFICATE_FILE = os.getenv ('PROXY_CERTIFICATE_FILE')


# We'll also create a few reusable strings for later interactivity.
string_constants = dict()
string_constants['CONFIRM'] = 'OK, are you sure you want to do this? This script does not have an "undo" feature.'
string_constants['CANCEL'] = 'OK. Operation canceled.'
string_constants['WORKING'] = 'Working...'
string_constants['COMPLETE'] = 'Operation complete.'
string_constants['NETWORK_SELECTED'] = 'Network selected.'
string_constants['NO_VALID_OPTIONS'] = 'There are no valid options. Please try again with an API key that has access to the appropriate resources.'

# Some of the parameters we'll work with are optional. This string defines what value will be put into a cell corresponding with a parameter that is not set on that rule.
string_constants['NOT_APPLICABLE'] = 'N/A'

# This script is interactive; user choices and data will be stored here.
user_choices = dict()
user_data = dict()


# Initialize the Meraki Dashboard connection.
if USE_PROXY == 1:
    dashboard = meraki.DashboardAPI(API_KEY, suppress_logging=True, certificate_path=PROXY_CERTIFICATE_FILE)  # Connection with proxy
else:
    dashboard = meraki.DashboardAPI(API_KEY, suppress_logging=True) # Connection without proxy


print("Number of organizations: ")
print( len(dashboard.organizations.getOrganizations()))


if len(dashboard.organizations.getOrganizations()) == 1 :
    MERAKI_ORGANIZATION_ID = dashboard.organizations.getOrganizations()[0]["id"]
    print(MERAKI_ORGANIZATION_ID)
else:
    MERAKI_ORGANIZATION_ID = os.getenv('MERAKI_ORGANIZATION_ID')




def printj(ugly_json_object):

     # The json.dumps() method converts a JSON object into human-friendly formatted text
    pretty_json_string = json.dumps(ugly_json_object, indent = 2, sort_keys = False)

    return print(pretty_json_string)

class UserChoice:

    'A re-usable CLI option prompt.'

    def __init__(self, options_list=[], subject_of_choice='available options', single_option_noun='option', id_parameter='id', name_parameter='name', action_verb='choose', no_valid_options_message=string_constants['NO_VALID_OPTIONS']):

        self.options_list = options_list # options_list is a list of dictionaries containing attributes id_parameter and name_parameter
        self.subject_of_choice = subject_of_choice # subject_of_choice is a string that names the subject of the user's choice. It is typically a plural noun.
        self.single_option_noun = single_option_noun # single_option_noun is a string that is a singular noun corresponding to the subject_of_choice
        self.id_parameter = id_parameter # id_parameter is a string that represents the name of the sub-parameter that serves as the ID value for the option in options_list. It should be a unique value for usability.
        self.name_parameter = name_parameter # name_paraemter is a string that represents the name of the sub-parameter that serves as the name value for the option in options_list. It does not need to be unique.
        self.action_verb = action_verb # action_verb is a string that represents the verb of the user's action. For example, to "choose"
        self.no_valid_options_message = no_valid_options_message # no_valid_options_message is a string that represents an error message if options_list is empty

        # Confirm there are options in the list
        if len(self.options_list):
            print(f'We found {len(self.options_list)} {self.subject_of_choice}:')

            # Label each option and show the user their choices.
            option_index = 0

            for option in self.options_list:
                print(f"{option_index}. {option[self.id_parameter]} with name {option[self.name_parameter]}")
                option_index+=1

            print(f'Which {self.single_option_noun} would you like to {self.action_verb}?')
            self.active_option = int(input(f'Choose 0-{option_index-1}:'))

            # Ask until the user provides valid input.
            while self.active_option not in list(range(option_index)):
                print(f'{self.active_option} is not a valid choice. Which {self.single_option_noun} would you like to {self.action_verb}?')
                self.active_option = int(input(f'Choose 0-{option_index-1}:'))

            print(f'Your {self.single_option_noun} is {self.options_list[self.active_option][self.name_parameter]}.')

            # Assign the class id and name vars to the chosen item's
            self.id = self.options_list[self.active_option][self.id_parameter]
            self.name = self.options_list[self.active_option][self.name_parameter]



def get_ssid_info(*, networkId):
    # Get the SSID Info
    ssid_info = dashboard.wireless.getNetworkWirelessSsids(networkId=networkId)

    return(ssid_info)

def create_workbook():

    # Create a workbook with the appropriate column headers and/or worksheets
    # First we'll create the workbook, then we'll design the worksheets. Just like in Excel, by default, the workbook has one worksheet, titled 'Sheet'.
    new_workbook = openpyxl.Workbook()

    # We'll specify that the active worksheet is our ssid_info_worksheet
    ssid_info_worksheet = new_workbook.active

    # Let's rename the worksheet from 'Sheet' to something more descriptive
    ssid_info_worksheet.title = 'SSID-Info'

    # Name the columns for the ssid_info_worksheet. Think of this like a single-line CSV:
    ssid_info_title_row_headers = (
        'SSID',
        'Visibility',
        'Access',
        'Security',
        'Password',
        'DNS',
        'FW Rule Plocy',
        'FW Rule Protocol',
        'FW Rule Destination',
        'FW Rule Port',
        'FW Rule Comment'
    )


    # Add that title row to the worksheet
    ssid_info_worksheet.append(ssid_info_title_row_headers)

    # Let's make the title row bold for easier reading
    for row in ssid_info_worksheet.iter_rows():
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)

    # Now let's do the same for the Switch Ports Info.
    # First, create a separate worksheet for the Switch Ports Info
    switch_ports_info_worksheet = new_workbook.create_sheet(title='Switch-Ports-Info')

    # Name the columns for the vpn_prefs_worksheet and custom_performance_classes_worksheet.
    switch_ports_info_title_row_headers = (
        'Sede',
        'Switch',
        'Puerto físico',
        'Negociación',
        'Tipo',
        'Descripción',
        'VLAN',
        'Voice VLAN'
    )

    # Add the title rows to the appropriate worksheets
    switch_ports_info_worksheet.append(switch_ports_info_title_row_headers)

    # Let's make those title rows bold, too
    for row in switch_ports_info_worksheet.iter_rows():
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)

    print(f'Created formatted workbook.')

    return(new_workbook)

def add_ssid_info_to_workbook(workbook, network_id):

    # We'll specify that the active worksheet is our ssid_info worksheet
    ssid_info_worksheet = workbook['SSID-Info']

    # We'll also count the number of SSID to help the user know that it's working.
    ssid_count = 0

    # Let's add all the SSID Info to the SSID-Info worksheet
    for ssid_info in user_data['ssid_info']:

        ssid_info_ssid_name = ssid_info['name']
        print (ssid_info_ssid_name)
        ssid_info_visibility = ssid_info['visible']
        try:
          ssid_info_access = ssid_info['encryptionMode']
        except:
          ssid_info_access = 'No aplica'
          print ('No hay encryptionMode en esta red.')
        try:
          ssid_info_security = ssid_info['wpaEncryptionMode']
        except:
          ssid_info_security = 'No aplica'
          print ('No hay WPA en esta red.')
        try:
          ssid_info_password = ssid_info['psk']
        except:
          ssid_info_password = 'No aplica'
          print ('No hay PSK en esta red.')
        try:
          ssid_info_dns = ssid_info['dnsRewrite']['dnsCustomNameservers']
        except:
          ssid_info_dns = 'No aplica'
          print ('No hay DNS forzado en esta red.')
        try:
          fw_wireless_rules = dashboard.wireless.getNetworkWirelessSsidFirewallL3FirewallRules(network_id, ssid_count)
          ssid_info_fw_rule_policy = fw_wireless_rules['rules']['policy']
          ssid_info_fw_rule_protocol = fw_wireless_rules['rules']['protocol']
          ssid_info_fw_rule_destination = fw_wireless_rules['rules']['destCidr']
          ssid_info_fw_rule_port = fw_wireless_rules['rules']['destPort']
          ssid_info_fw_rule_comment = fw_wireless_rules['rules']['comment']
        except:
          ssid_info_fw_rule_policy = 'No aplica'
          ssid_info_fw_rule_protocol = 'No aplica'
          ssid_info_fw_rule_destination = 'No aplica'
          ssid_info_fw_rule_comment = 'No aplica'
          print ('No existen reglas de firewall definidas para este SSID')
        # We assemble the parameters into a tuple that will represent a row.
        ssid_info_row = (
            str(ssid_info_ssid_name),
            str(ssid_info_visibility),
            str(ssid_info_access),
            str(ssid_info_security),
            str(ssid_info_password),
            str(ssid_info_dns),
            str(ssid_info_fw_rule_policy),
            str(ssid_info_fw_rule_protocol),
            str(ssid_info_fw_rule_destination),
            str(ssid_info_fw_rule_comment)
            )

        # We then add that row to the worksheet.
        ssid_info_worksheet.append(ssid_info_row)

        # increase the ssid_count
        ssid_count += 1


    print(f'Added {ssid_count} performance classes to customPerformanceClasses.')

    return(workbook)


# Function that saves a workbook
def save_workbook(workbook):

    # Finally, we save the worksheet.
    # Let's give it a name with a date/time stamp
    downloaded_rules_workbook_filename = f'ssid_info_workbook_{datetime.datetime.now()}.xlsx'.replace(':','')

    workbook.save(downloaded_rules_workbook_filename)

    print(f'Saved {downloaded_rules_workbook_filename}.')

# Function that load a workbook template to read values to set/configure network.
def load_excel(excel_name):
    try:
        print ('Loading excel...')
        excel_document = openpyxl.load_workbook(excel_name)
        print(excel_document)
        print(excel_document.sheetnames)
        print ('Excel loaded! :)')

        return excel_document

    except Exception as e:
        print("Error loading excel file...")
        print(e)
        return (e)






app = Flask(__name__)
port = "5000"
bootstrap = Bootstrap(app)
app.secret_key = API_KEY
app.config['SESSION_TYPE'] = 'filesystem'


# Define Flask routes
@app.route("/")
def index():
    user_choices['all_organizations'] = dashboard.organizations.getOrganizations()
    info_organization = dashboard.organizations.getOrganization(MERAKI_ORGANIZATION_ID)
    info_organization_clients = dashboard.organizations.getOrganizationClientsOverview(MERAKI_ORGANIZATION_ID)
    info_networks = dashboard.organizations.getOrganizationNetworks(MERAKI_ORGANIZATION_ID, total_pages='all')
    network_index = random.randint(0, len(info_networks)-1)
    print(len(info_networks))
    print(network_index)
    print(info_networks[network_index]['id'])
    info_alerts = dashboard.networks.getNetworkAlertsHistory(info_networks[network_index]['id'], total_pages='all')

    return render_template('index.html', info_organization_v=info_organization['name'], info_networks_v=info_networks[network_index]['name'], info_alerts_v=len(info_alerts), info_organization_clients_up_v=round(info_organization_clients["usage"]["overall"]["upstream"]/1000, 2), info_organization_clients_down_v=round(info_organization_clients["usage"]["overall"]["downstream"]/1000, 2))

@app.route("/job-order")
def job_order():
    return render_template('job-order.html')

@app.route("/add-GP-FW-Rules")
def add_GP_FW_Rules():
    return render_template('add-GP-FW-Rules.html')

@app.route('/add-GP-FW-Rules-read', methods = ['POST'])
def add_GP_FW_Rules_read():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        JOB_ORDER_EXCEL_NAME = f.filename
        print (JOB_ORDER_EXCEL_NAME)


        job_order = load_excel(JOB_ORDER_EXCEL_NAME)
        #df = pd.DataFrame(job_order.values)
        df = pd.read_excel(
            JOB_ORDER_EXCEL_NAME,    # name of excel sheet
            names=['ID','Policy', 'Protocol', 'Destination', 'Port', 'Comment'],       # new column header
            skiprows=5,   # list of rows you want to omit at the beginning
            #index_col=0,
            skipfooter=3,         # number of rows you want to skip at the end
            usecols='A:F'       # columns to parse (note the excel-like syntax)
        )
        df.to_csv('GP-FW-Rules-to-add.csv', encoding='utf-8')
        info_network = job_order['GROUP POLICY']['B3'] # Sede
        info_GP = job_order['GROUP POLICY']['B4'] # Sede
        session['info_network'] = info_network.value
        session['info_GP'] = info_GP.value

        return render_template("add-GP-FW-Rules-read.html", filename=JOB_ORDER_EXCEL_NAME, info_network_v=info_network.value, info_GP_v=info_GP.value,  column_names=df.columns.values, row_data=list(df.values.tolist()), link_column="ID", zip=zip)

@app.route('/add-GP-FW-Rules-config')
def add_GP_FW_Rules_config():
    df = pd.read_csv('GP-FW-Rules-to-add.csv', encoding='utf8')
    print("Given Dataframe :\n", df)

    info_networks = dashboard.organizations.getOrganizationNetworks(MERAKI_ORGANIZATION_ID, total_pages='all')
    info_network = session.get('info_network', None)
    print(info_network)
    print(info_networks)
    print(len(info_networks))
    print(info_networks[0]['name'])
    try:
        for info_network_index in range(len(info_networks)):
            if info_networks[info_network_index]['name'] == info_network:
                print(info_network_index)
                info_network_id = info_networks[info_network_index]['id']
                print(info_network_id)
    except ValueError:
        print(f"Lo sentimos, no se ha podido encontrar la sede {info_network} en la organización")
    info_GPs = dashboard.networks.getNetworkGroupPolicies(info_network_id)
    info_GP = session.get('info_GP', None)
    print(info_GP)
    try:
        for info_GP_index in range(len(info_GPs)):
            if info_GPs[info_GP_index]['name'] == info_GP:
                print(info_GP_index)
                info_GP_id = info_GPs[info_GP_index]['groupPolicyId']
                print(info_GP_id)

    except ValueError:
        print(f"Lo sentimos, no se ha podido encontrar la política de grupo {info_GP} en la sede {info_network}")
    info_GP_object = dashboard.networks.getNetworkGroupPolicy( info_network_id, info_GP_id)

    info_GP_FW_Rule_comment = 'Allow TCP traffic to subnet with HTTP servers.'
    info_GP_FW_Rule_policy = 'allow'
    info_GP_FW_Rule_protocol = 'tcp'
    info_GP_FW_Rule_destPort = '443'
    info_GP_FW_Rule_destCIDR = '10.0.0.0/24'

    info_GP_FW_Rules_List = []

    print("Given Dataframe :\n", df)
    for ind in df.index:
        print(df['ID'][ind], df['Policy'][ind], df['Protocol'][ind], df['Destination'][ind], df['Port'][ind], df['Comment'][ind])
        info_FW_Rule = {'comment': df['Comment'][ind], 'policy': df['Policy'][ind], 'protocol': df['Protocol'][ind], 'destPort': df['Port'][ind], 'destCidr': df['Destination'][ind]}
        info_GP_FW_Rules_List.append(info_FW_Rule)

    info_firewallAndTrafficShaping = {'l3FirewallRules': info_GP_FW_Rules_List }

    try:
        response = dashboard.networks.updateNetworkGroupPolicy(info_network_id, info_GP_id, firewallAndTrafficShaping=info_firewallAndTrafficShaping )
        message = '¡Configuración aplicada con éxito!'
    except APIError as e:
        print(f"Lo sentimos, ha ocurrido un error al intentar cargar la regla: {info_GP_FW_Rule_policy} | {info_GP_FW_Rule_protocol} | {info_GP_FW_Rule_destCIDR} | {info_GP_FW_Rule_destPort} | {info_GP_FW_Rule_comment} en la sede {info_network}")
        if e.status == 200:
            message = '¡Configuración aplicada con éxito!' 
        else:
            print(e.status)
            print(e.reason)
            print(e.message)
            message = 'Ha ocurrido un error al intentar cargar la configuración, inténtalo de nuevo'   

    return render_template("add-GP-FW-Rules-configured.html")

@app.route("/clear-GP-FW-Rules")
def clear_GP_FW_Rules():
    return render_template('clear-GP-FW-Rules.html')

@app.route("/add-SSID-FW-Rules")
def add_SSID_FW_Rules():
    return render_template('add-GP-FW-Rules.html')

@app.route("/clear-SSID-FW-Rules")
def clear_SSID_FW_Rules():
    return render_template('clear-GP-FW-Rules.html')

@app.route("/about")
def about():
    with open("requirements.txt", "r") as f:
        content = f.read()
    return render_template('about.html', content=content)



# Start the Flask server in a new thread
threading.Thread(target=app.run, kwargs={"use_reloader": False, "debug": True}).start()